import re
import unicodedata
import datetime as dt

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\ludwi\OneDrive\Escritorio\kj\Cruzimex\Analisis_Trade"
DATA = BASE + r"\Data"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def norm(s):
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s


def pareto_abc(values: pd.Series) -> pd.Series:
    """Standard 80/15/5 pareto classification.
    The item that pushes the cumulative total past a threshold is still
    classified using the cumulative % BEFORE it is added (inclusive rule)."""
    s = values.sort_values(ascending=False)
    total = s.sum()
    if total == 0:
        return pd.Series("C", index=values.index)
    cum = s.cumsum() / total
    prev_cum = cum.shift(1, fill_value=0)

    def classify(pc):
        if pc < 0.80:
            return "A"
        elif pc < 0.95:
            return "B"
        else:
            return "C"

    abc = prev_cum.apply(classify)
    return abc.reindex(values.index)


def timedelta_to_str(td):
    if pd.isna(td):
        return None
    total_seconds = int(td.total_seconds())
    h, rem = divmod(total_seconds, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def mode_timedelta(series: pd.Series):
    series = series.dropna()
    series = series[series > dt.timedelta(0)]
    if series.empty:
        return None
    m = series.mode()
    if m.empty:
        return None
    return m.iloc[0]


def floor_to_minute(td):
    total_seconds = int(td.total_seconds())
    return dt.timedelta(minutes=total_seconds // 60)


def timedelta_to_hhmm(td):
    if pd.isna(td) or td is None:
        return None
    total_minutes = int(td.total_seconds() // 60)
    h, m = divmod(total_minutes, 60)
    return f"{h:02d}:{m:02d}"


def moda_tiempo_diario(rec_df, group_col):
    """Promedio diario de tiempo atendido (sin ceros), truncado a minutos,
    y luego la moda de esos promedios diarios por grupo."""
    base = rec_df[rec_df["tiempo_td"] > dt.timedelta(0)].copy()
    daily_avg = (
        base.groupby([group_col, base["fecha"].dt.date])["tiempo_td"]
        .mean()
        .reset_index(name="promedio_dia")
    )
    daily_avg["promedio_dia_min"] = daily_avg["promedio_dia"].apply(floor_to_minute)
    return daily_avg.groupby(group_col)["promedio_dia_min"].apply(mode_timedelta)


# ---------------------------------------------------------------------------
# 1. Load clientes.xlsx
# ---------------------------------------------------------------------------

clientes = pd.read_excel(DATA + r"\clientes.xlsx", sheet_name="Sheet1")
clientes = clientes.rename(columns={
    "Codigo Cliente": "cod_cliente",
    "Nombre cliente": "nombre_cliente",
    "Razon Social": "razon_social",
    "Clasificacion": "clasificacion_cliente",
    "Ruta asignada": "ruta",
})
clientes["cod_cliente"] = clientes["cod_cliente"].astype(str).str.strip()
clientes["norm_nombre"] = clientes["nombre_cliente"].apply(norm)
clientes["ruta"] = clientes["ruta"].astype(str).str.strip()

# Solo analizamos rutas de bodega (WHS) + SC-LICORES RAMADA, excluyendo explicitamente
# SC-WHS-MERC-LA-RAMADA-2
RUTA_EXCLUIDA = "SC-WHS-MERC-LA-RAMADA-2"


def ruta_incluida(r):
    if r == RUTA_EXCLUIDA:
        return False
    return ("WHS" in r) or (r == "SC-LICORES RAMADA")


n_antes = len(clientes)
clientes = clientes[clientes["ruta"].apply(ruta_incluida)].copy()
print(f"Filtro de rutas (WHS + SC-LICORES RAMADA, excluye {RUTA_EXCLUIDA}): {n_antes} -> {len(clientes)} codigos")

# Diccionario de unificacion de rutas -> Mercado. Las rutas que solo difieren por
# el sufijo numerico de sub-ruta (-1, -2, -3, -4) se agrupan en un unico mercado.
# Los numeros que son parte del nombre real del mercado (KM-6, PLAN-3000,
# 4-DE-NOVIEMBRE, 1ERO-DE-MARZO/MAYO) NO se recortan. SC-LICORES RAMADA queda
# tal cual, sin unificarse a ningun mercado.
RUTA_A_MERCADO = {
    "SC-WHS-MERC-1ERO-DE-MARZO": "MERCADO 1RO DE MARZO",
    "SC-WHS-MERC-4-DE-NOVIEMBRE": "MERCADO 4 DE NOVIEMBRE",
    "SC-WHS-MERC-ABASTO-1": "MERCADO ABASTO",
    "SC-WHS-MERC-ABASTO-2": "MERCADO ABASTO",
    "SC-WHS-MERC-ALTO-SAN-PEDRO": "MERCADO ALTO SAN PEDRO",
    "SC-WHS-MERC-BELEN": "MERCADO BELEN",
    "SC-WHS-MERC-CENTRAL-DE-WARNES": "MERCADO CENTRAL DE WARNES",
    "SC-WHS-MERC-EL-FUERTE": "MERCADO EL FUERTE",
    "SC-WHS-MERC-EL-TROMPILLO": "MERCADO EL TROMPILLO",
    "SC-WHS-MERC-ESTACION-ARGENTINA": "MERCADO ESTACION ARGENTINA",
    "SC-WHS-MERC-FERIA-BARRIO-LINDO": "MERCADO FERIA BARRIO LINDO",
    "SC-WHS-MERC-FORTALEZA": "MERCADO FORTALEZA",
    "SC-WHS-MERC-GUAPURU": "MERCADO GUAPURU",
    "SC-WHS-MERC-KM-6": "MERCADO KM 6",
    "SC-WHS-MERC-LA-CAMPANA": "MERCADO LA CAMPANA",
    "SC-WHS-MERC-LA-CUCHILLA": "MERCADO LA CUCHILLA",
    "SC-WHS-MERC-LA-GUARDIA": "MERCADO LA GUARDIA",
    "SC-WHS-MERC-LA-PAMPITA": "MERCADO LA PAMPITA",
    "SC-WHS-MERC-LA-RAMADA-1": "MERCADO LA RAMADA",
    "SC-WHS-MERC-LA-RAMADA-3": "MERCADO LA RAMADA",
    "SC-WHS-MERC-LA-RAMADA-4": "MERCADO LA RAMADA",
    "SC-WHS-MERC-LOS-BOSQUES": "MERCADO LOS BOSQUES",
    "SC-WHS-MERC-LOS-POCITOS": "MERCADO LOS POCITOS",
    "SC-WHS-MERC-LOS-POZOS-1": "MERCADO LOS POZOS",
    "SC-WHS-MERC-LOS-POZOS-2": "MERCADO LOS POZOS",
    "SC-WHS-MERC-MUTUALISTA-1": "MERCADO MUTUALISTA",
    "SC-WHS-MERC-MUTUALISTA-2": "MERCADO MUTUALISTA",
    "SC-WHS-MERC-NUEVO-ABASTO": "MERCADO NUEVO ABASTO",
    "SC-WHS-MERC-NUEVO-ABASTO-2": "MERCADO NUEVO ABASTO",
    "SC-WHS-MERC-NUEVO-LA-RAMADA": "MERCADO NUEVO LA RAMADA",
    "SC-WHS-MERC-NUEVO-LOS-POZOS": "MERCADO NUEVO LOS POZOS",
    "SC-WHS-MERC-NUEVO-PLAN-3000": "MERCADO NUEVO PLAN 3000",
    "SC-WHS-MERC-NUEVO-PLAN-3000-2": "MERCADO NUEVO PLAN 3000",
    "SC-WHS-MERC-PLAN-3000": "MERCADO PLAN 3000",
    "SC-WHS-MERC-PRIMAVERA": "MERCADO PRIMAVERA",
    "SC-WHS-MERC-RAMAFA": "MERCADO RAMAFA",
    "SC-WHS-MERC-SAN-JUAN": "MERCADO SAN JUAN",
    "SC-WHS-MERC-SAN-LUIS": "MERCADO SAN LUIS",
    "SC-WHS-MERC-SANTA-CRUZ-SUR": "MERCADO SANTA CRUZ SUR",
    "SC-WHS-MERC-SATELITE-NORTE": "MERCADO SATELITE NORTE",
    "SC-WHS-MERC-VILLA-1ERO-DE-MAYO": "MERCADO VILLA 1RO DE MAYO",
    "SC-WHS-MERC-VILLA-1ERO-DE-MAYO-2": "MERCADO VILLA 1RO DE MAYO",
    "SC-WHS-MONTERO-MERC-ABASTO": "MERCADO MONTERO ABASTO",
    "SC-WHS-MONTERO-MERC-GERMAN-MORENO": "MERCADO MONTERO GERMAN MORENO",
    "SC-LICORES RAMADA": "SC-LICORES RAMADA",  # queda tal cual, sin unificar
}

rutas_sin_mapear = sorted(set(clientes["ruta"]) - set(RUTA_A_MERCADO))
if rutas_sin_mapear:
    print("AVISO: rutas sin mapeo en el diccionario (se dejan tal cual):", rutas_sin_mapear)

clientes["ruta"] = clientes["ruta"].map(lambda r: RUTA_A_MERCADO.get(r, r))
print(f"Rutas unificadas a {clientes['ruta'].nunique()} mercados")

# Build merged "entity" per (norm_nombre, ruta) to collapse duplicate-name clients
grp_key = ["norm_nombre", "ruta"]
entity_ids = clientes.groupby(grp_key).ngroup()
clientes["entity_id"] = entity_ids

entity_master = (
    clientes.groupby("entity_id")
    .agg(
        cod_cliente=("cod_cliente", lambda x: "/".join(sorted(x))),
        nombre_cliente=("nombre_cliente", "first"),
        ruta=("ruta", "first"),
        norm_nombre=("norm_nombre", "first"),
        raw_codes=("cod_cliente", lambda x: list(x)),
        n_codigos=("cod_cliente", "count"),
    )
    .reset_index()
)

code_to_entity = dict(zip(clientes["cod_cliente"], clientes["entity_id"]))
code_to_ruta = dict(zip(clientes["cod_cliente"], clientes["ruta"]))

n_merged = (entity_master["n_codigos"] > 1).sum()
print(f"Clientes.xlsx: {len(clientes)} codigos -> {len(entity_master)} entidades ({n_merged} fusionadas)")

# ---------------------------------------------------------------------------
# 2. Load sales.xlsx and restrict to Jan-Jul 2026, clients present in clientes.xlsx
# ---------------------------------------------------------------------------

sales = pd.read_excel(DATA + r"\sales.xlsx", sheet_name="Ventas Combo Armado")
sales["Fecha_dt"] = pd.to_datetime(sales["Fecha"], format="%d/%m/%Y", errors="coerce")
mask_periodo = (sales["Fecha_dt"] >= "2026-01-01") & (sales["Fecha_dt"] <= "2026-07-31")
sales = sales[mask_periodo].copy()

sales["Cliente"] = sales["Cliente"].astype(str).str.strip()
sales["entity_id"] = sales["Cliente"].map(code_to_entity)
sales["ruta_cliente"] = sales["Cliente"].map(code_to_ruta)

sales_valid = sales.dropna(subset=["entity_id"]).copy()
sales_valid["entity_id"] = sales_valid["entity_id"].astype(int)

print(f"sales.xlsx: {len(sales)} filas en el periodo, {len(sales_valid)} cruzan con clientes.xlsx")

# --- Route level sales totals ---
ruta_ventas = sales_valid.groupby("ruta_cliente")["Venta Neta"].sum()

# --- Entity level sales totals + sku count ---
entity_ventas = sales_valid.groupby("entity_id")["Venta Neta"].sum()
entity_skus = sales_valid.groupby("entity_id")["Articulo"].nunique()

# ---------------------------------------------------------------------------
# 3. Load clientes_reccorridos.xlsx
# ---------------------------------------------------------------------------

rec = pd.read_excel(DATA + r"\clientes_reccorridos.xlsx", sheet_name="Sheet1")
rec = rec.rename(columns={
    "Fecha": "fecha",
    "Reponedor": "reponedor",
    "Cliente": "nombre_cliente_rec",
    "Ruta": "ruta",
    "Inicio de Reposición": "inicio",
    "Fin de Reposición": "fin",
    "Tiempo Atendido": "tiempo_atendido",
    "Estado": "estado",
})
rec["norm_nombre"] = rec["nombre_cliente_rec"].apply(norm)
rec["ruta"] = rec["ruta"].astype(str).str.strip()
rec["ruta"] = rec["ruta"].map(lambda r: RUTA_A_MERCADO.get(r, r))
rec["fecha"] = pd.to_datetime(rec["fecha"])

# Map to entity primarily by name (clientes.xlsx route is authoritative and can
# differ from the route logged in recorridos for the same real client). Only use
# the recorridos route as a tie-breaker when a name maps to more than one entity
# (i.e. the same full name legitimately belongs to different clients on
# different routes).
name_to_entities = {}
for _, er in entity_master.iterrows():
    name_to_entities.setdefault(er["norm_nombre"], []).append((er["ruta"], er["entity_id"]))


def resolve_entity(norm_nombre, ruta):
    candidates = name_to_entities.get(norm_nombre)
    if not candidates:
        return np.nan
    if len(candidates) == 1:
        return candidates[0][1]
    matches = [eid for r, eid in candidates if r == ruta]
    if len(matches) == 1:
        return matches[0]
    return np.nan  # ambiguous: same name, multiple routes, none matches exactly


rec["entity_id"] = [resolve_entity(n, r) for n, r in zip(rec["norm_nombre"], rec["ruta"])]

unmatched = rec[rec["entity_id"].isna()]
print(f"clientes_reccorridos.xlsx: {len(rec)} filas, {len(unmatched)} sin cruce con clientes.xlsx (excluidas)")
if len(unmatched):
    print("  nombres sin cruce (muestra):", sorted(unmatched['nombre_cliente_rec'].unique())[:20])
    print("  total nombres distintos sin cruce:", unmatched['nombre_cliente_rec'].nunique())

rec_valid = rec.dropna(subset=["entity_id"]).copy()
rec_valid["entity_id"] = rec_valid["entity_id"].astype(int)

# Parse tiempo_atendido -> timedelta
def to_timedelta(v):
    if isinstance(v, dt.time):
        return dt.timedelta(hours=v.hour, minutes=v.minute, seconds=v.second)
    if isinstance(v, dt.timedelta):
        return v
    if isinstance(v, str):
        try:
            h, m, s = [int(x) for x in v.split(":")]
            return dt.timedelta(hours=h, minutes=m, seconds=s)
        except Exception:
            return pd.NaT
    return pd.NaT

rec_valid["tiempo_td"] = rec_valid["tiempo_atendido"].apply(to_timedelta)

ATENDIDO_ESTADOS = {"VISITADO", "INCOMPLETO"}
rec_valid["atendido"] = rec_valid["estado"].isin(ATENDIDO_ESTADOS)

entity_to_ruta = dict(zip(entity_master["entity_id"], entity_master["ruta"]))
rec_valid["ruta_oficial"] = rec_valid["entity_id"].map(entity_to_ruta)

rec_atendidos = rec_valid[rec_valid["atendido"]].copy()

# ---------------------------------------------------------------------------
# 4. Hoja1: ReposicionxRuta
# ---------------------------------------------------------------------------

rutas_all = sorted(entity_master["ruta"].unique())

clientes_por_ruta = entity_master.groupby("ruta")["entity_id"].nunique()

atendidos_por_ruta = rec_atendidos.groupby("ruta_oficial")["entity_id"].nunique()

# Tiempo atendido por ruta: promedio diario (HH:MM) de tiempo atendido, y luego
# la moda de esos promedios diarios a lo largo del mes.
moda_tiempo_ruta = moda_tiempo_diario(rec_atendidos, "ruta_oficial")

# Media clientes atendidos por dia: para cada ruta, contar clientes distintos atendidos por dia,
# luego promedio (min) y percentil75 (max) de esos conteos diarios
daily_counts = (
    rec_atendidos.groupby(["ruta_oficial", rec_atendidos["fecha"].dt.date])["entity_id"]
    .nunique()
    .reset_index(name="clientes_dia")
)
media_por_ruta = daily_counts.groupby("ruta_oficial")["clientes_dia"].mean()
p75_por_ruta = daily_counts.groupby("ruta_oficial")["clientes_dia"].quantile(0.75)

hoja1_rows = []
for ruta in rutas_all:
    total_clientes = int(clientes_por_ruta.get(ruta, 0))
    total_atendidos = int(atendidos_por_ruta.get(ruta, 0))
    sin_atencion = total_clientes - total_atendidos
    pct_atencion = (total_atendidos / total_clientes) if total_clientes else 0
    total_ventas = float(ruta_ventas.get(ruta, 0.0))
    tiempo_moda = moda_tiempo_ruta.get(ruta)
    media_val = media_por_ruta.get(ruta)
    p75_val = p75_por_ruta.get(ruta)
    if pd.notna(media_val) and pd.notna(p75_val):
        media_str = f"{media_val:.0f}-{p75_val:.0f}"
    else:
        media_str = None

    hoja1_rows.append({
        "Ruta": ruta,
        "Clasificacion ABC (de la ruta)": None,  # se calcula luego con pareto
        "Total ventas Ruta": total_ventas,
        "Cantidad de clientes de esa ruta": total_clientes,
        "Cantidad de clientes atendidos": total_atendidos,
        "Clientes sin Atencion": sin_atencion,
        "% Atencion": pct_atencion,
        "Tiempo Atendido (en base a la moda)": timedelta_to_hhmm(tiempo_moda) if tiempo_moda is not None else None,
        "Media clientes atendidos por Dia": media_str,
    })

hoja1 = pd.DataFrame(hoja1_rows)
abc_ruta = pareto_abc(hoja1.set_index("Ruta")["Total ventas Ruta"])
hoja1["Clasificacion ABC (de la ruta)"] = hoja1["Ruta"].map(abc_ruta)
hoja1 = hoja1.sort_values("Total ventas Ruta", ascending=False).reset_index(drop=True)

ruta_abc_map = dict(zip(hoja1["Ruta"], hoja1["Clasificacion ABC (de la ruta)"]))

# ---------------------------------------------------------------------------
# 5. Global ABC per entity: ventas & skus (universo = todos los clientes en clientes.xlsx)
# ---------------------------------------------------------------------------

all_entity_ids = entity_master["entity_id"]
ventas_full = entity_ventas.reindex(all_entity_ids, fill_value=0.0)
skus_full = entity_skus.reindex(all_entity_ids, fill_value=0)

abc_ventas_global = pareto_abc(ventas_full)
abc_skus_global = pareto_abc(skus_full)

# Visitas / moda de tiempo / atendido flag por entidad
visitas_por_entidad = rec_atendidos.groupby("entity_id").size()
# Tiempo atendido por cliente: promedio diario (HH:MM), luego moda de esos promedios
moda_tiempo_entidad = moda_tiempo_diario(rec_atendidos, "entity_id")
entidades_atendidas = set(rec_atendidos["entity_id"].unique())

# ---------------------------------------------------------------------------
# 6. Hoja2: ReposicionxCliente (solo atendidos)
# ---------------------------------------------------------------------------

hoja2_rows = []
hoja3_rows = []

for _, row in entity_master.iterrows():
    eid = row["entity_id"]
    ruta = row["ruta"]
    total_vendido = float(ventas_full.get(eid, 0.0))
    n_skus = int(skus_full.get(eid, 0))
    abc_v = abc_ventas_global.get(eid, "C")
    abc_s = abc_skus_global.get(eid, "C")
    base = {
        "Ruta": ruta,
        "ABC_Ruta": ruta_abc_map.get(ruta),
        "Cod.Cliente": row["cod_cliente"],
        "Nombre Cliente": row["nombre_cliente"],
        "Total vendido BS.": total_vendido,
        "ABC_Ventas": abc_v,
        "Cantidad de SKUs": n_skus,
        "ABC_SKUs": abc_s,
    }
    if eid in entidades_atendidas:
        moda_t = moda_tiempo_entidad.get(eid)
        base2 = dict(base)
        base2["Moda de tiempo"] = timedelta_to_hhmm(moda_t) if moda_t is not None else None
        base2["Cantidad de Visitas"] = int(visitas_por_entidad.get(eid, 0))
        hoja2_rows.append(base2)
    else:
        hoja3_rows.append(base)

hoja2 = pd.DataFrame(hoja2_rows).sort_values(["Ruta", "Total vendido BS."], ascending=[True, False]).reset_index(drop=True)
hoja3 = pd.DataFrame(hoja3_rows).sort_values(["Ruta", "Total vendido BS."], ascending=[True, False]).reset_index(drop=True)

print(f"Hoja2 (atendidos): {len(hoja2)} clientes")
print(f"Hoja3 (sin reposicion): {len(hoja3)} clientes")

# ---------------------------------------------------------------------------
# 7. Write formatted workbook
# ---------------------------------------------------------------------------

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ABC_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="FFC7CE"),
}

MONEY_COLS = {"Total ventas Ruta", "Total vendido BS."}
PCT_COLS = {"% Atencion"}
ABC_COLS = {"Clasificacion ABC (de la ruta)", "ABC_Ruta", "ABC_Ventas", "ABC_SKUs"}


def write_sheet(name, df):
    ws = wb.create_sheet(name)
    ws.freeze_panes = "A2"
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            val = r[col]
            if isinstance(val, float) and np.isnan(val):
                val = None
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            if col in MONEY_COLS:
                c.number_format = '#,##0.00'
            elif col in PCT_COLS:
                c.number_format = '0.0%'
            elif col in ABC_COLS and val in ABC_FILL:
                c.fill = ABC_FILL[val]
                c.alignment = Alignment(horizontal="center")

    for j, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:500]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 3, 12), 45)

    ws.auto_filter.ref = ws.dimensions
    return ws


write_sheet("ReposicionxRuta", hoja1)
write_sheet("ReposicionxCliente", hoja2)
write_sheet("ClientesSinReposicion", hoja3)

timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
OUT = BASE + rf"\SC_Analisis_TradeMarketing_{timestamp}.xlsx"
wb.save(OUT)
print("Guardado:", OUT)
